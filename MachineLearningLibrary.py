import os.path
import shutil
import numpy as np
import numpy.matlib
from imageio import imread, imsave
from PIL import Image
import imageio
import scipy
import scipy.stats
import shelve
import sklearn
import sklearn.preprocessing
import sklearn.svm
import sklearn.decomposition
import sklearn.calibration
import sklearn.metrics
import sklearn.feature_selection
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.feature_selection import SelectKBest, f_classif
import skimage.feature
import skimage.color
import skimage.transform
from skimage import draw
from skimage.transform import resize as imresize
from skimage.util import img_as_ubyte
import xlsxwriter
from openpyxl.utils import get_column_letter, column_index_from_string
import matplotlib.pyplot as plt

class ML(object):

    # Initialized all variables describing the folder structure
    def __init__(self):

        # Folders
        self.num = 1
        self.resizedImageFolder = 'StandardizedData'
        self.lbpFolder = 'LocalBinaryPatterns'
        self.vwMatricesFolder_CrossValidation = 'LDAMatrices_CrossValidation'
        self.vwMatricesFolder = 'LDAMatrices'
        self.featuresFolder_CrossValidation = 'Features_CrossValidation'
        self.featuresFolder = 'Features'
        self.landmarksFolder = 'Points_GT'
        self.resultsFolder = 'Results'

        # Files
        self.crossValidationModelName = 'CrossValidation'
        self.crossValidationExcelName = 'CrossValidation.xlsx'

        self.finalLDAMatricesName = 'LDAMatrices'
        self.finalModelName = 'Model'
        self.finalModelExcelName = 'FinalModelStatistics.xlsx'


        # Class names
        self.negativeClassName = 'Normal'
        self.positiveClassName = 'Syndromic'

        # Configuration
        self.featureSelector = 'RecursiveElimination' # options = 'MCFS_p' or 'RecursiveElimination'

        self.LBP_list = np.array([[16,4],[32,8],[64,12]]) # LBP resolutions [elements, radius]
        self.Ws = 5 # Windows size for the LBP calculation

        self.numberOfLandmarks = 44

        self.threshold = 0.5
        self.optimizeThreshold = False

        self.verbose = False

        self.asymmetry = False

        self.finalNumberOfFeatures = 30

    #################################################################################
    ## THE FUNCTIONS BELOW ARE THE CLASS INTERFACE (PUBLIC ACCESS)
    #################################################################################

    # This will 1) Create the directory tree 2) Create and save standardized images and landmarks, 3) Save local binary patterns
    def Initialize(self, negativeClassImageFolder, negativeClassLandmarkFolder, positiveClassImageFolder, positiveClassLandmarkFolder, workingPath, removeExistingFiles=False):

        # creating the folder structure
        self.createOutputFolderStructure(workingPath, removeExistingFiles=removeExistingFiles)

        # Standardizing the negative class
        self.rescaleImagesAlignLandmarks(negativeClassImageFolder, negativeClassLandmarkFolder, os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName))

        # Standardizing the positive class
        self.rescaleImagesAlignLandmarks(positiveClassImageFolder, positiveClassLandmarkFolder, os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName))

        # Calculating and saving the local binary patterns
        self.SaveLocalBinaryPatterns(workingPath)

    # Saves all features for cross validations
    def CalculateFeaturesForCrossValidation(self, workingPath):

        self.CalculateLDAMatrices_CrossValidation(workingPath)

        negativeImageFolder = os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName)
        negativeLandmarkFolder = os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName, self.landmarksFolder)
        positiveImageFolder = os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName)
        positiveLandmarkFolder = os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName, self.landmarksFolder)

        listOfImages = []
        landmarks = np.ndarray([0, self.numberOfLandmarks, 2], dtype=np.float32)

        img_list = sorted(os.listdir(negativeImageFolder))
        for f_i in range(len(img_list)):

            fileName, fileExtension = os.path.splitext(img_list[f_i])

            if fileExtension.lower() == '.jpg':

                # Adding the image to the list
                image = imread(os.path.join(negativeImageFolder, img_list[f_i]))

                # Adding the landmarks
                landmarks = np.load(os.path.join(negativeLandmarkFolder, fileName + '.npy'))[:self.numberOfLandmarks, :]

                # Read V and W matrices for this patient
                vWDictionary = {}
                shelveDictionary = shelve.open(os.path.join(workingPath, self.vwMatricesFolder_CrossValidation, self.negativeClassName+'_'+fileName))
                for key, value in shelveDictionary.items():
                    vWDictionary[key] = value
                shelveDictionary.close()

                features, description = self.GetFeatures([image], landmarks.reshape(1,landmarks.shape[0], landmarks.shape[1]), vWDictionary)

                featureDictionary = shelve.open(os.path.join(workingPath, self.featuresFolder_CrossValidation, self.negativeClassName+'_'+fileName))
                featureDictionary['features'] = features
                featureDictionary['description'] = description
                featureDictionary.close()

        img_list = sorted(os.listdir(positiveImageFolder))
        for f_i in range(len(img_list)):

            fileName, fileExtension = os.path.splitext(img_list[f_i])

            if fileExtension.lower() == '.jpg':

                # Adding the image to the list
                image = imread(os.path.join(positiveImageFolder, img_list[f_i]))

                # Adding the landmarks
                landmarks = np.load(os.path.join(positiveLandmarkFolder, fileName + '.npy'))[:self.numberOfLandmarks, :]

                # Read V and W matrices for this patient
                vWDictionary = {}
                shelveDictionary = shelve.open(os.path.join(workingPath, self.vwMatricesFolder_CrossValidation, self.positiveClassName+'_'+fileName))
                for key, value in shelveDictionary.items():
                    vWDictionary[key] = value
                shelveDictionary.close()

                features, description = self.GetFeatures([image], landmarks.reshape(1,landmarks.shape[0], landmarks.shape[1]), vWDictionary)

                featureDictionary = shelve.open(os.path.join(workingPath, self.featuresFolder_CrossValidation, self.positiveClassName+'_'+fileName))
                featureDictionary['features'] = features
                featureDictionary['description'] = description
                featureDictionary.close()

    # Leave-one-out cross validation. The results are saved in the working directory (outputFolder)
    def CrossValidate(self, workingPath, maxNumberOfFeatures=30):

        # geometricsPath: folder containing the geometric features
        # lbpPath: folder containing the local binary patterns
        # vwPath: folder containing the V and W matrices to create the texture features from the local binary patterns (different matrices for each case left out)
        # savePath: folder to save the cross validation results
        # maxNumberOfFeatures: maximum number of features to cross validate

        ## Load features
        featurePath = os.path.join(workingPath, self.featuresFolder_CrossValidation)
        savePath = os.path.join(workingPath, self.resultsFolder, self.crossValidationModelName)

        patientList = []

        img_list = sorted(os.listdir(featurePath))
        for f_i in range(len(img_list)):

            fileName, fileExtension = os.path.splitext(img_list[f_i])

            # Load features
            shelveDictionary = shelve.open(os.path.join(featurePath, fileName))
            patientFeatures = shelveDictionary['features']

            if 'featureDescription' not in locals():
                featureDescription = shelveDictionary['description']
            shelveDictionary.close()

            if not self.asymmetry:
                patientFeatures_new = []
                featureDescription_new = []
                for i in range(len(featureDescription)):
                    if "Asymmetry" not in featureDescription[i]:
                        patientFeatures_new += [patientFeatures[0,i]]
                        featureDescription_new += [featureDescription[i]]

                new_num = len(patientFeatures_new)
                patientFeatures = np.array(patientFeatures_new)
                patientFeatures = np.reshape(patientFeatures, (1,new_num))


            # Getting if the patient is normal or syndromic, and his name/ID
            parts = fileName.split('_')
            type = parts[0]
            patientName = parts[1]

            patientList += [patientName]


            if type==self.negativeClassName:
                patientLabel = 0
            else:
                patientLabel = 1

            if 'allLabels' not in locals():
                allLabels = np.array([patientLabel], np.int)
                allFeatures = patientFeatures

            else:
                allLabels = np.concatenate([allLabels, np.array([patientLabel], dtype=np.int)], axis=0)
                allFeatures = np.concatenate([allFeatures, patientFeatures], axis=0)

        ###################################################################################
        #allFeatures = allFeatures[:,:18] # keeping only the geometric features!
        #featureDescription = featureDescription[:18]
        ###################################################################################

        nFeatures = allFeatures.shape[1]
        nImages = allFeatures.shape[0]

        # We sort so we first have all normals and then all syndromics
        sortIndices = np.argsort(allLabels)
        allLabels = allLabels[sortIndices]
        allFeatures = allFeatures[sortIndices,:]
        patientList = [patientList[i] for i in sortIndices]


        # Statistics on the features
        negativeClassMean = np.mean(allFeatures[np.where(allLabels==0)[0],:], axis=0).ravel()
        negativeClassStd = np.std(allFeatures[np.where(allLabels==0)[0],:], axis=0).ravel()
        positiveClassMean = np.mean(allFeatures[np.where(allLabels==1)[0],:], axis=0).ravel()
        positiveClassStd = np.std(allFeatures[np.where(allLabels==1)[0],:], axis=0).ravel()

        pValues = np.zeros([nFeatures], dtype=np.float)
        for i in range(nFeatures):
            try:
                _, pValues[i] = scipy.stats.mannwhitneyu(allFeatures[np.where(allLabels==0)[0],i], allFeatures[np.where(allLabels==1)[0],i], alternative='two-sided')
            except:
                pValues[i] = 1

        statistics = (negativeClassMean, negativeClassStd, positiveClassMean, positiveClassStd, pValues)

        # Maximum number of features to try
        dim = np.min([nFeatures, nImages, maxNumberOfFeatures])

        # Probabilities estimated
        probs = np.zeros([nImages, dim], dtype=np.float32)
        rocArea = np.zeros([dim], dtype=np.float32)
        sensitivity = np.zeros([dim], dtype=np.float32)
        specificity = np.zeros([dim], dtype=np.float32)
        accuracy = np.zeros([dim], dtype=np.float32)
        threshold = np.zeros([dim], dtype=np.float32)
        f1 = np.zeros([dim], dtype=np.float32)
        selectedFeatures = np.zeros([nFeatures, dim], dtype=np.bool) # Matrix of selected features (True is selected, False otherwise)

        for k in range(dim):
            if self.verbose:
                print('Number of features: {}'.format(str(k+1)))

            # Scaling data
            scaler = sklearn.preprocessing.StandardScaler()
            scaledData = scaler.fit_transform(allFeatures)

            # Selecting the k+1 most important features
            if self.featureSelector == 'MCFS_p':
                options = {}
                options['gnd'] = np.zeros(allLabels.shape)
                options['gnd'][np.where(allLabels==0)[0]] = 1
                options['gnd'][np.where(allLabels==1)[0]] = 2
                featureIds = MCFS_p(scaledData, k+1, options)[0].ravel()
            elif self.featureSelector == 'RecursiveElimination':
                model = sklearn.svm.SVC(kernel='linear', probability=True, class_weight='balanced')
                selector = sklearn.feature_selection.RFE(model, k+1, step=1)
                selector.fit(scaledData, allLabels)
                featureIds = np.argwhere(selector.ranking_==1).ravel()
            elif self.featureSelector == 'SequentialSelection':
                model = sklearn.svm.SVC(kernel='linear', probability=True, class_weight='balanced')
                selector = SequentialFeatureSelector(model, n_features_to_select=k+1, direction="forward")
                selector.fit(scaledData, allLabels)
                featureIds = np.argwhere(selector.get_support()==1).ravel()
                # print(featureIds)
            elif self.featureSelector == 'SelectKBest':
                selector = SelectKBest(f_classif, k=50)
                selector.fit(scaledData, allLabels)
            else:
                raise Exception('Unknown feature selector')

            featureIds = np.sort(featureIds)

            # Updating the selected features matrix
            selectedFeatures[featureIds,k] = True

            # Leave one out
            for testId in range(nImages):

                train = np.ones([allLabels.size], dtype=bool)
                test = np.zeros([allLabels.size], dtype=bool)

                train[testId] = False
                test[testId] = True

                trainClasses = allLabels[train]
                testClasses = allLabels[test]

                trainData = scaledData[train, :]
                trainData = trainData[:, featureIds]
                if trainData.ndim == 1: # It means there is only one feature
                    trainData = np.reshape(trainData, [trainData.size, 1])

                testData = scaledData[test, :]
                testData = testData[:, featureIds]
                if testData.ndim == 1: # It means there is only one feature
                    testData = np.reshape(testData, [testData.size, 1])

                model = sklearn.svm.SVC(kernel='linear', probability=True, class_weight='balanced')
                model.fit(trainData, trainClasses)
                probs[testId, k] = model.predict_proba(testData)[0, 1] # probability of the second class (syndromic)

            rocArea[k] = sklearn.metrics.roc_auc_score(allLabels, probs[:, k])

            if self.optimizeThreshold:

                fpr, tpr, thresh = sklearn.metrics.roc_curve(allLabels, probs[:, k], pos_label=1, drop_intermediate=False)


                optIndex = np.argmax(tpr-fpr)
                threshold[k] = thresh[optIndex]
                self.threshold = threshold[k]

                sensitivity[k] = tpr[optIndex]
                specificity[k] = 1-fpr[optIndex]
                accuracy[k] = np.mean((probs[:, k]  >= threshold[k]) == allLabels.astype(np.bool))

                if k==5:
                    print(accuracy[k])
                    excelPath = os.path.join(workingPath, self.resultsFolder, 'Prediction_CV.xlsx')
                    workbook = xlsxwriter.Workbook(excelPath)
                    worksheet = workbook.add_worksheet('Prediction')
                    worksheet.write(0, 0, 'GT')
                    worksheet.write(0, 1, 'Leave 1 out')
                    for i in range(len(probs)):
                        worksheet.write(i+1, 0, allLabels[i])
                        worksheet.write(i+1, 1, probs[:, k][i] >= threshold[k])
                    workbook.close()

                tp = np.sum(np.logical_and((probs[:, k] >= threshold[k]), (allLabels == 1)))
                fp = np.sum(np.logical_and((probs[:, k] >= threshold[k]), (allLabels == 0)))
                tn = np.sum(np.logical_and((probs[:, k] < threshold[k]), (allLabels == 0)))
                fn = np.sum(np.logical_and((probs[:, k] < threshold[k]), (allLabels == 1)))
                precision = tp/(tp+fp)
                recall = sensitivity[k]
                f1[k] = 2 * (precision * recall)/ (precision + recall)

            else:

                threshold[k] = self.threshold


                tp = np.sum(np.logical_and((probs[:, k] >= threshold[k]), (allLabels == 1)))
                fp = np.sum(np.logical_and((probs[:, k] >= threshold[k]), (allLabels == 0)))
                tn = np.sum(np.logical_and((probs[:, k] < threshold[k]), (allLabels == 0)))
                fn = np.sum(np.logical_and((probs[:, k] < threshold[k]), (allLabels == 1)))

                sensitivity[k] = tp/(tp+fn)
                specificity[k] = tn/(tn+fp)
                accuracy[k] = (tp+tn)/(tp+fp+tn+fn)

            if self.verbose:
                #print('  Feature list: ' + str(listOfFeatures))
                print('  AUC: {:.2f}. Accuracy: {:.2f}. Sensitivity: {:.2f}. Specificity: {:.2f}. Threshold: {:.2f}'.format(rocArea[k], accuracy[k], sensitivity[k], specificity[k], threshold[k]))

        shelveDictionary = shelve.open(savePath)
        shelveDictionary['patientList'] = patientList
        shelveDictionary['labels'] = allLabels
        shelveDictionary['probs'] = probs
        shelveDictionary['rocArea'] = rocArea
        shelveDictionary['sensitivity'] = sensitivity
        shelveDictionary['specificity'] = specificity
        shelveDictionary['accuracy'] = accuracy
        shelveDictionary['threshold'] = threshold
        shelveDictionary['selectedFeatures'] = selectedFeatures
        if not self.asymmetry:
            featureDescription = featureDescription_new
        shelveDictionary['featureDescription'] = featureDescription
        shelveDictionary['statistics'] = statistics
        shelveDictionary['configuration'] = vars(self) # Saving the configuration
        shelveDictionary['f1'] = f1
        shelveDictionary['allFeatures'] = allFeatures
        shelveDictionary.close()

    # Export cross validation results to an Excel file
    def ExportCrossValidationToExcel(self, workingPath):

        crossValidationPath = os.path.join(workingPath, self.resultsFolder, self.crossValidationModelName)
        excelPath = os.path.join(workingPath, self.resultsFolder, self.crossValidationExcelName)

        shelveDictionary = shelve.open(crossValidationPath)
        patientList = shelveDictionary['patientList']
        allLabels = shelveDictionary['labels']
        probs = shelveDictionary['probs']
        rocArea = shelveDictionary['rocArea']
        sensitivity = shelveDictionary['sensitivity']
        specificity = shelveDictionary['specificity']
        accuracy = shelveDictionary['accuracy']
        threshold = shelveDictionary['threshold']
        selectedFeatures = shelveDictionary['selectedFeatures']
        featureDescription = shelveDictionary['featureDescription']
        f1 = shelveDictionary['f1']
        allFeatures = shelveDictionary['allFeatures']
        (negativeClassMean, negativeClassStd, positiveClassMean, positiveClassStd, pValues) = shelveDictionary['statistics']
        shelveDictionary.close()

        # Create a workbook
        workbook = xlsxwriter.Workbook(excelPath)
        # Formats
        boldFormat = workbook.add_format({'bold': True, 'align': 'center'})
        decimalFormat = workbook.add_format({'num_format': '0.00', 'align': 'center'})
        scientificFormat = workbook.add_format({'num_format': '0.00E+00', 'align': 'center'})

        # Sheet 1: Statistics of features
        worksheet = workbook.add_worksheet('AllFeatures')
        worksheet.write(0, 0, 'Feature', boldFormat)
        worksheet.write(0, 1, 'Mean: '+self.negativeClassName, boldFormat)
        worksheet.write(0, 2, 'Std: '+self.negativeClassName, boldFormat)
        worksheet.write(0, 3, 'Mean: '+self.positiveClassName, boldFormat)
        worksheet.write(0, 4, 'Std: '+self.positiveClassName, boldFormat)
        worksheet.write(0, 5, 'p-value', boldFormat)
        for i in range(len(featureDescription)):
            worksheet.write(i+1, 0, featureDescription[i])
            worksheet.write(i+1, 1, negativeClassMean[i], decimalFormat)
            worksheet.write(i+1, 2, negativeClassStd[i], decimalFormat)
            worksheet.write(i+1, 3, positiveClassMean[i], decimalFormat)
            worksheet.write(i+1, 4, positiveClassStd[i], decimalFormat)
            worksheet.write(i+1, 5, pValues[i], scientificFormat)

        # Sheet 2: Cross validation results
        worksheet = workbook.add_worksheet('CrossValidation')
        worksheet.write(0, 0, '# features', boldFormat)
        worksheet.write(0, 1, 'ROC area', boldFormat)
        worksheet.write(0, 2, 'Accuracy', boldFormat)
        worksheet.write(0, 3, 'Sensitivity', boldFormat)
        worksheet.write(0, 4, 'Specificity', boldFormat)
        worksheet.write(0, 5, 'Threshold', boldFormat)
        worksheet.write(0, 6, 'F1 score', boldFormat)
        for i in range(accuracy.size):
            worksheet.write(i+1, 0, i+1)
            worksheet.write(i+1, 1, rocArea[i], decimalFormat)
            worksheet.write(i+1, 2, accuracy[i], decimalFormat)
            worksheet.write(i+1, 3, sensitivity[i], decimalFormat)
            worksheet.write(i+1, 4, specificity[i], decimalFormat)
            worksheet.write(i+1, 5, threshold[i], decimalFormat)
            worksheet.write(i+1, 6, f1[i], decimalFormat)

        chart = workbook.add_chart({'type': 'line'})

        chart.add_series({  'values': ['CrossValidation', 1, 1, accuracy.size, 1], #[sheetname, first_row, first_col, last_row, last_col]
                            'categories': ['CrossValidation', 1, 0, accuracy.size, 0],
                            'line': {'color': 'navy', 'dash_type': 'solid', 'width': 2.5},
                            'name': 'ROC area'})

        chart.add_series({  'values': ['CrossValidation', 1, 2, accuracy.size, 2], #[sheetname, first_row, first_col, last_row, last_col]
                            'categories': ['CrossValidation', 1, 0, accuracy.size, 0],
                            'line': {'color': 'red', 'dash_type': 'solid', 'width': 2.5},
                            'name': 'Accuracy'})

        chart.add_series({  'values': ['CrossValidation', 1, 3, accuracy.size, 3], #[sheetname, first_row, first_col, last_row, last_col]
                            'categories': ['CrossValidation', 1, 0, accuracy.size, 0],
                            'line': {'color': 'purple', 'dash_type': 'solid', 'width': 1.25},
                            'name': 'Sensitivity'})

        chart.add_series({  'values': ['CrossValidation', 1, 4, accuracy.size, 4], #[sheetname, first_row, first_col, last_row, last_col]
                            'categories': ['CrossValidation', 1, 0, accuracy.size, 0],
                            'line': {'color': 'green', 'dash_type': 'solid', 'width': 1.25},
                            'name': 'Specificity'})

        chart.add_series({  'values': ['CrossValidation', 1, 5, accuracy.size, 5], #[sheetname, first_row, first_col, last_row, last_col]
                            'categories': ['CrossValidation', 1, 0, accuracy.size, 0],
                            'line': {'color': 'yellow', 'dash_type': 'dash', 'width': 1.00},
                            'name': 'Threshold'})

        chart.set_x_axis({'name': '# features'})
        chart.set_y_axis({'name': '# features', 'min': 0, 'max': 1.05})

        worksheet.insert_chart('G1', chart)

        # Sheet 3: Selected features
        worksheet = workbook.add_worksheet('Selectedfeatures')
        worksheet.write(0, 0, 'Feature', boldFormat) # Headers

        for i in range(selectedFeatures.shape[1]):
            worksheet.write(0, i+1, str(i+1), boldFormat)
        worksheet.write(0, selectedFeatures.shape[1]+1, 'SUM', boldFormat)

        for i in range(len(featureDescription)): # Data
            #print(featureDescription[i])
            worksheet.write(i+1, 0, featureDescription[i])
            for j in range(selectedFeatures.shape[1]):
                worksheet.write(i+1, j+1, selectedFeatures[i,j])
        for i in range(len(featureDescription)): # Formula
            worksheet.write_formula(i+1, selectedFeatures.shape[1]+1, '=SUM('+get_column_letter(2)+str(i+2)+':'+get_column_letter(selectedFeatures.shape[1]+1)+str(i+2)+')', boldFormat)

        # Sheet 4: All feature values
        worksheet = workbook.add_worksheet('FeatureValues')
        worksheet.write(0, 0, 'Feature', boldFormat) # Headers
        print(allFeatures.shape)

        for i in range(len(featureDescription)):
            worksheet.write(i+1, 0, featureDescription[i])
        for i in range(len(patientList)):
            worksheet.write(0, i+1, patientList[i])
            for j in range(len(featureDescription)):
                worksheet.write(j+1, i+1, allFeatures[i,j])

        workbook.close()

    ##################################################################################################
    ## THE FUNCTIONS BELOW DO NOT HAVE TO BE USED OUTSIDE THE CLASS UNLESS YOU KNOW WHAT YOU ARE DOING
    ##################################################################################################

    # Creates the folder structure within the working (output) folder
    def createOutputFolderStructure(self, outputFolder, removeExistingFiles=False):

        if removeExistingFiles and os.path.exists(outputFolder):
            shutil.rmtree(outputFolder)

        if not os.path.exists(os.path.join(outputFolder, self.resizedImageFolder, self.negativeClassName, self.landmarksFolder)):
            os.makedirs(os.path.join(outputFolder, self.resizedImageFolder, self.negativeClassName, self.landmarksFolder))

        if not os.path.exists(os.path.join(outputFolder, self.resizedImageFolder, self.positiveClassName, self.landmarksFolder)):
            os.makedirs(os.path.join(outputFolder, self.resizedImageFolder, self.positiveClassName, self.landmarksFolder))

        if not os.path.exists(os.path.join(outputFolder, self.lbpFolder)):
            os.makedirs(os.path.join(outputFolder, self.lbpFolder))

        if not os.path.exists(os.path.join(outputFolder, self.vwMatricesFolder_CrossValidation)):
            os.makedirs(os.path.join(outputFolder, self.vwMatricesFolder_CrossValidation))

        # if not os.path.exists(os.path.join(outputFolder, self.vwMatricesFolder)):
        #     os.makedirs(os.path.join(outputFolder, self.vwMatricesFolder))

        if not os.path.exists(os.path.join(outputFolder, self.featuresFolder_CrossValidation)):
            os.makedirs(os.path.join(outputFolder, self.featuresFolder_CrossValidation))

        # if not os.path.exists(os.path.join(outputFolder, self.featuresFolder)):
        #     os.makedirs(os.path.join(outputFolder, self.featuresFolder))

        if not os.path.exists(os.path.join(outputFolder, self.resultsFolder)):
            os.makedirs(os.path.join(outputFolder, self.resultsFolder))

    # Standardize the images and landmarks for one case
    def normalizeImageAndLandmarks(self, image, landmarks, interPupilDist = 150, orientation = 0, sizeOut = 150, offsetOut = 11):
        # image: the image to process
        # landmarks: the landmarks to process, shape [nLandmarks, 2]
        # sizeOut: size of the output image. (150)
        # offsetOut: offset to both sides (rows and columns) in the transformed image
        # interPupilDist: distance between pupil landmarks (5 and 10)
        # orientation: orientation of the interpupil vector


        interPupilVector = landmarks[9, :] - landmarks[4, :]

        Or_i, Mag_i = cart2pol(interPupilVector[0], interPupilVector[1]) # Cartesian coordinates of the vector

        scalingFactor = interPupilDist / Mag_i

        Or_i = -(orientation - Or_i) # using (-) sign b/c Y axis in image space is inverted

        center_ldmks = np.array([np.mean(landmarks[:,0]), np.mean(landmarks[:,1]), 1.0], dtype=np.float32)

        # Transformation matrix
        T_mat = np.array([[scalingFactor*np.cos(Or_i),-scalingFactor *np.sin(Or_i),0],[scalingFactor*np.sin(Or_i), scalingFactor*np.cos(Or_i),0],[0,0,1]],dtype=np.float32)
        invT_mat = np.linalg.inv(T_mat) # inverse

        landmarks = np.append(landmarks, np.ones([landmarks.shape[0], 1]), axis=1)

        # Transformation of landmarks (rotation and scaling)
        P_align = np.zeros(landmarks.shape, dtype=np.float32)
        for i in range(landmarks.shape[0]):
            P_align[i, :] = np.dot( landmarks[i,:]-center_ldmks, T_mat) + center_ldmks

        ##########################################################
        # Now everything is aligned using T_mat wrt center_ldmks
        ##########################################################

        #  Box around the inner landmarks
        colRange = np.array([np.floor(min(P_align[:33,0])), np.ceil(max(P_align[:33,0]))]).astype(np.int)
        rowRange = np.array([np.floor(min(P_align[:33,1])), np.ceil(max(P_align[:33,1]))]).astype(np.int)

        colSize = colRange[1] - colRange[0]
        rowSize = rowRange[1] - rowRange[0]
        totalLength = max(colSize, rowSize) # This is the size of the squared box around the inner landmarks

        colOffset = int(np.ceil( (offsetOut * (colSize + 1) )/ (sizeOut - 2 * offsetOut) ) )
        rowOffset = int(np.ceil( (offsetOut * (rowSize + 1) )/ (sizeOut - 2 * offsetOut) ) )
        offset = max(colOffset, rowOffset) # Offset to add (in the new space of aligned landmarks)

        # Extending the ranges to fit the squared box
        if colSize <= rowSize:
            colRange = np.array([colRange[0]-np.round((rowSize-colSize)/2), colRange[1]+np.round((rowSize-colSize)/2)]).astype(np.int)
        else:
            rowRange = np.array([rowRange[0]-np.round((colSize-rowSize)/2), rowRange[1]+np.round((colSize-rowSize)/2)]).astype(np.int)

        # This are the ranges of the aligned landmarks space to keep
        colRange_new = np.array([(colRange[0] - offset), (colRange[1] + offset)])
        rowRange_new = np.array([(rowRange[0] - offset), (rowRange[1] + offset)])

        # The new image
        newImg = np.zeros([rowRange_new[1]-rowRange_new[0] + 1, colRange_new[1]-colRange_new[0] + 1], dtype=np.uint8)

        for newX in range(colRange_new[0], colRange_new[1] + 1):

            for newY in range(rowRange_new[0], rowRange_new[1] + 1):

                # Invert previous transformation to set the interpupilary distance
                oCoords = np.dot(np.array([newX, newY, 1.0], dtype=np.float32) - center_ldmks, invT_mat) + center_ldmks

                if oCoords[0] < image.shape[1] and oCoords[1] < image.shape[0] and oCoords[0]>=0 and oCoords[1]>=0:

                    newImg[newY-rowRange_new[0], newX-colRange_new[0]] = image[int(oCoords[1]), int(oCoords[0])]


        # Applying the offset to the landmarks
        P_align[:,0] -= colRange_new[0]
        P_align[:,1] -= rowRange_new[0]

        # Scaling image and landmarks
        scalingFactor = sizeOut/newImg.shape[1]
        P_align[:,0] *= scalingFactor
        scalingFactor = sizeOut/newImg.shape[0]
        P_align[:,1] *= scalingFactor
        newImg = skimage.transform.resize(newImg, (sizeOut, sizeOut))

        return newImg, P_align


    # Rescales and aligns all images in a data folder
    def rescaleImagesAlignLandmarks(self, imPath, landmarkPath, sPath, interPupilDist = 150, orientation = 0, sizeOut = 150, offsetOut = 10):
        # imPath -> Folder containing all images. Assumes all images are within the same folder
        # sPath-> Folder to save all aligned and rescaled images
        # interPupilDist: distance between pupil landmarks (5 and 10)
        # orientation: orientation of the interpupil vector
        # sizeOut: size of the output image. (150)
        # offsetOut: offset to both sides (rows and columns) in the transformed image

        img_list = sorted(os.listdir(imPath))

        for file in img_list:

            fileName, fileExtension = os.path.splitext(file)

            if fileExtension.lower() == '.jpg':
            # if fileExtension.lower() == '.png':
                if self.verbose:
                    print('Processing: ' + file)

                # Grayscale reading
                rgb_weights = [0.2989, 0.5870, 0.1140]
                I = np.dot(imread(os.path.join(imPath,file))[..., :3], rgb_weights)

                # Read pts files to get landmarks
                P = np.loadtxt(os.path.join(landmarkPath, fileName + '.pts'), unpack = True).T

                # Normalizing
                newImg, P_align  = self.normalizeImageAndLandmarks(I, P, interPupilDist = interPupilDist, orientation = orientation, sizeOut = sizeOut, offsetOut = offsetOut)

                ##################################################################
                # 4. Save Scaled and Cropped Images/Landmarks
                ##################################################################
                imageio.imsave(os.path.join(sPath, fileName + '.jpg'), img_as_ubyte(newImg))
                np.save(os.path.join(sPath, self.landmarksFolder, fileName), P_align[:,:2])

    # Returns the local binary patterns for one case
    def extractLBP(self, image, landmarks, LBP_list, Ws):

        patientDictionary = {}

        patientDictionary['FeatureDescriptions'] = ['Texture at columella',
                                'Texture at nasion',
                                'Texture at tip of nose',
                                'Texture at philtrum',
                                'Texture at center of cupid\'s bow',
                                'Texture at lower border of upper lip',
                                'Texture at upper border of lower lip',
                                'Average texture at lateral canthi',
                                'Average texture at lower eyelids',
                                'Average texture at medial canthi',
                                'Average texture at upper eyelids',
                                'Average texture at center of the pupil',
                                'Average texture at lateral of nose root',
                                'Average texture at alar crease',
                                'Average texture at center of ala',
                                'Average texture at bottom of ala',
                                'Average texture at nostril top', #alar septum junction
                                'Average texture at oral commissures',
                                'Average texture at side of cupid\'s bow',
                                'Average texture at side of lower lip',
                                'Asymmetry of texture at lateral canthi',
                                'Asymmetry of texture at lower eyelids',
                                'Asymmetry of texture at medial canthi',
                                'Asymmetry of texture at upper eyelids',
                                'Asymmetry of texture at center of the eyes',
                                'Asymmetry of texture at lateral of nose root',
                                'Asymmetry of texture at alar crease',
                                'Asymmetry of texture at center of ala',
                                'Asymmetry of texture at bottom of ala',
                                'Asymmetry of texture at nostril top',
                                'Asymmetry of texture at oral commissures',
                                'Asymmetry of texture at side of cupid\'s bow',
                                'Asymmetry of texture at side of lower lip']

        centerIndices = np.array([15, 21, 22, 23, 26, 31, 32], dtype=np.int)
        leftIndices = np.array([0, 1, 2, 3, 4, 10, 11, 12, 13, 14, 24, 25, 30], dtype=np.int)
        rightIndices = np.array([7, 6, 5, 8, 9, 20, 19, 18, 17, 16, 28, 27, 29], dtype=np.int)

        # Normalizing image values between 0 and 1 --> equivalent to MATLAB mat2gray function
        min_ = np.min(image)
        max_ = np.max(image)
        I_norm = (image - min_) / (max_ - min_ + np.spacing(1))

        # Expand the image to avoid accessing to the wrong location in case
        # the filter on the LBP patch were too big.
        P_Extra = np.max(np.array([LBP_list[0,1],LBP_list[1,1],LBP_list[2,1],Ws])) + 5

        I_Corner_UL = I_norm[(P_Extra-1)::-1,(P_Extra-1)::-1]
        I_Corner_UR = I_norm[(P_Extra-1)::-1,(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1]
        I_Corner_BL = I_norm[(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1,(P_Extra-1)::-1]
        I_Corner_BR = I_norm[(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1,(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1]
        I_u = I_norm[(P_Extra-1)::-1,:]
        I_b = I_norm[(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1,:]
        I_l = I_norm[:,(P_Extra-1)::-1]
        I_r = I_norm[:,(I_norm.shape[1]-1):(I_norm.shape[1]-1)-(P_Extra+2):-1]

        I_Ext_left = np.vstack((np.hstack((I_Corner_UL,I_u,I_Corner_UR)),np.hstack((I_l,I_norm,I_r)),np.hstack((I_Corner_BL,I_b,I_Corner_BR))))

        (Nr,Nc) = I_Ext_left.shape

        # Create image filter mask
        Mid_r = np.linspace((-((Ws-1)/2)),((Ws-1)/2),int(Ws)).reshape(1,int(Ws))
        Mid_c = Mid_r.T
        Mid_r = Nr*Mid_r
        W_IDs = np.matlib.repmat(Mid_r,Ws,1) + np.matlib.repmat(Mid_c,1,Ws)
        W_IDs_vector = np.reshape(W_IDs,(W_IDs.size,1, 1))

        x_align_left = landmarks[:33, 0].reshape(1, 33) + P_Extra # This is for left and center indices
        y_align_left = landmarks[:33, 1].reshape(1, 33) + P_Extra

        # Flipping the image and landmarks
        I_Ext_right = np.fliplr(I_Ext_left)
        centerX = I_Ext_left.shape[1]/2.0
        x_align_right = (x_align_left - centerX)*(-1) + centerX # For right indices
        y_align_right = np.copy(y_align_left)

        # Pass Landmarks to IDs
        Ldmks_IDs_center = np.zeros([centerIndices.size], dtype=np.int)
        Ldmks_IDs_left = np.zeros([leftIndices.size], dtype=np.int)
        Ldmks_IDs_right = np.zeros([rightIndices.size], dtype=np.int)
        for i in range(centerIndices.size):
            Ldmks_IDs_center[i] = np.ravel_multi_index((y_align_left[0,centerIndices[i]].astype(int), x_align_left[0,centerIndices[i]].astype(int)), I_Ext_left.shape,mode='raise')
        for i in range(leftIndices.size):
            Ldmks_IDs_left[i] = np.ravel_multi_index((y_align_left[0,leftIndices[i]].astype(int), x_align_left[0,leftIndices[i]].astype(int)), I_Ext_left.shape,mode='raise')
        for i in range(rightIndices.size):
            Ldmks_IDs_right[i] = np.ravel_multi_index((y_align_right[0,rightIndices[i]].astype(int), x_align_right[0,rightIndices[i]].astype(int)), I_Ext_left.shape,mode='raise')

        Ldmks_IDs_center = np.reshape(Ldmks_IDs_center,(1,1,Ldmks_IDs_center.size))
        Ldmks_IDs_left = np.reshape(Ldmks_IDs_left,(1,1,Ldmks_IDs_left.size))
        Ldmks_IDs_right = np.reshape(Ldmks_IDs_right,(1,1,Ldmks_IDs_right.size))

        # LBP LOOP ---> IN
        ##################################################################
        for l_i in range(LBP_list.shape[0]): # For each LBP resolution: pair [element,radius]

            LBP_i = LBP_list[l_i,:].reshape(1,LBP_list.shape[1]) # LBP_i = (# elements x radii)

            # Create LBP Pattern
            LBP_Pattern = generateRadialFilterLBP(LBP_i[0,0],LBP_i[0,1]) # output is W x W x #elements, where W is 2 x radius + 1

            # Create ID-Mask for the Neighborwood LBP patch
            W_Ngh = LBP_Pattern.shape[0]
            Mid_r = np.arange(-(W_Ngh-1)/2, (W_Ngh-1)/2 + 1).astype(np.int).reshape(1,int(W_Ngh))
            Mid_c = Mid_r.T
            Mid_r = Nr*Mid_r
            Ngh_IDs_aux = np.matlib.repmat(Mid_r,W_Ngh,1) +  np.matlib.repmat(Mid_c,1,W_Ngh)

            # LBP_diff_vect LOOP ---> IN
            ##############################################################
            for p_i in range(LBP_Pattern.shape[2]): # For each element in the neighborhood (there are #element elements)

                LBP_ij = np.copy(LBP_Pattern[:,:,p_i]) # LBP_ij is the WxW window -> As of here, it is independent from the coordinates of the landmarks

                # Find neighbors
                IDs = np.where(LBP_ij > 0) # Pixels in the window LBP_ij to compare with
                IDs_offset = np.insert(np.reshape(Ngh_IDs_aux[IDs], [1,IDs[0].size]), 0, 0, axis=1)

                # Find weights
                W_lbp = np.insert(np.reshape(LBP_ij[IDs], [1,IDs[0].size]), 0, -1, axis=1)

                #Compute the total offset
                W_IDs_MAT = np.tile(W_IDs_vector, [1, IDs_offset.size, x_align_left.size])


                LBP_ij_IDs_MAT = np.tile(np.reshape(IDs_offset, [IDs_offset.shape[0], IDs_offset.shape[1], 1]),[W_IDs_vector.size, 1, x_align_left.size])
                W_lbp_MAT = np.tile(np.reshape(W_lbp, [W_lbp.shape[0], W_lbp.shape[1], 1]), [W_IDs_vector.size, 1, x_align_left.size])

                IDs_MAT = W_IDs_MAT + LBP_ij_IDs_MAT

                Ldmks_IDs_MAT_center = np.tile(Ldmks_IDs_center, [IDs_MAT.shape[0], IDs_MAT.shape[1], 1])
                Ldmks_IDs_MAT_left = np.tile(Ldmks_IDs_left, [IDs_MAT.shape[0], IDs_MAT.shape[1], 1])
                Ldmks_IDs_MAT_right = np.tile(Ldmks_IDs_right, [IDs_MAT.shape[0], IDs_MAT.shape[1], 1])

                All_IDs_center = Ldmks_IDs_MAT_center + IDs_MAT[:,:,centerIndices]
                All_IDs_left = Ldmks_IDs_MAT_left + IDs_MAT[:,:,leftIndices]
                All_IDs_right = Ldmks_IDs_MAT_right + IDs_MAT[:,:,rightIndices]

                I_all_IDs_int_center = I_Ext_left.ravel()[All_IDs_center.ravel().astype(np.int)].reshape(All_IDs_center.shape)
                I_all_IDs_int_left = I_Ext_left.ravel()[All_IDs_left.ravel().astype(np.int)].reshape(All_IDs_left.shape)
                I_all_IDs_int_right = I_Ext_right.ravel()[All_IDs_right.ravel().astype(np.int)].reshape(All_IDs_right.shape)

                I_all_IDs_int_center = np.sum(I_all_IDs_int_center * W_lbp_MAT[:,:,centerIndices], axis=1, keepdims=True)
                I_all_IDs_int_left = np.sum(I_all_IDs_int_left * W_lbp_MAT[:,:, leftIndices], axis=1, keepdims=True)
                I_all_IDs_int_right = np.sum(I_all_IDs_int_right * W_lbp_MAT[:,:,rightIndices], axis=1, keepdims=True)

                # Appending center and side features (average and difference)
                I_all_IDs_int = np.concatenate([I_all_IDs_int_center, (I_all_IDs_int_left+I_all_IDs_int_right)/2, np.abs(I_all_IDs_int_left-I_all_IDs_int_right)], axis=2)

                if p_i == 0:
                    LBP_All_i = np.copy(I_all_IDs_int)
                else:
                    LBP_All_i = np.append(LBP_All_i, I_all_IDs_int, axis=1)


            patientDictionary['LBP_'+str(l_i + 1)] = LBP_All_i

        return patientDictionary

    # Calculates the local binary patterns for all images in imagePath, with landmarks under imagePath/self.landmarksFolder
    def SaveLocalBinaryPatterns(self, workingPath):
        # imagePath: folder with all the images. The landmarks are under the folder self.landmarksFolder in that directory
        # outputPath: folder to save the calculated local binary patterns
        # Ws: Window size

        outputPath = os.path.join(workingPath, self.lbpFolder)

        # Doing it for the negative class
        imagePath = os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName)
        landmarksPath = os.path.join(imagePath, self.landmarksFolder)

        img_list = sorted(os.listdir(imagePath))
        for file in img_list:

            fileName, fileExtension = os.path.splitext(file)

            if fileExtension == '.jpg':

                patientDictionary = {}

                # Read image
                image = imread(os.path.join(imagePath,file))

                # Read landmarks
                landmarks = np.load(os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName, self.landmarksFolder, fileName+'.npy'))

                # Calculate LBP
                lbpDictionary = self.extractLBP(image, landmarks, self.LBP_list, self.Ws)

                # Saving patient LBPs
                shelveDictionary = shelve.open(os.path.join(outputPath, self.negativeClassName + '_' + fileName))
                for key, value in lbpDictionary.items():

                    if key == "FeatureDescriptions":

                        new_value = []
                        new_key = []
                        for i in value:

                           # print("Asymmetry" in str(i))
                            if "Asymmetry" not in str(i):
                                new_value.append(i)
                                #value.remove(i)
                        value = new_value
                         #else:
                          #
                            # print(shelveDictionary[key])
                    shelveDictionary[key] = value
                    #print(shelveDictionary[key])


                #print(shelveDictionary.keys())
                shelveDictionary.close()


        # Doing it for the positive class
        imagePath = os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName)
        landmarksPath = os.path.join(imagePath, self.landmarksFolder)

        img_list = sorted(os.listdir(imagePath))
        for file in img_list:

            fileName, fileExtension = os.path.splitext(file)

            if fileExtension == '.jpg':

                patientDictionary = {}

                # Read image
                image = imread(os.path.join(imagePath,file))

                # Read landmarks
                landmarks = np.load(os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName, self.landmarksFolder, fileName+'.npy'))

                # Calculate LBP
                lbpDictionary = self.extractLBP(image, landmarks, self.LBP_list, self.Ws)

                # Saving patient LBPs
                shelveDictionary = shelve.open(os.path.join(outputPath, self.positiveClassName + '_' + fileName))
                for key, value in lbpDictionary.items():
                    if key == "FeatureDescriptions":

                        new_value = []
                        new_key = []
                        for i in value:

                            # print("Asymmetry" in str(i))
                            if "Asymmetry" not in str(i):
                                new_value.append(i)
                                # value.remove(i)
                        value = new_value
                    shelveDictionary[key] = value
                shelveDictionary.close()


    # Creates the W and V matrices used to calculate the texture features from the local binary patterns.
    # Since it is used for cross validation, it calculates different matrices for each subject
    # Assumes that the local binary patterns are already calculated in their folder
    def CalculateLDAMatrices_CrossValidation(self, workingPath):

        lbpPath = os.path.join(workingPath, self.lbpFolder)
        outputPath = os.path.join(workingPath, self.vwMatricesFolder_CrossValidation)

        T = 20 # number of iterations
        d1r = 1 # the size of the window filter will be d1 x d1r (where d1 is the original size of the patch)
        d2r = 1 # the size of the window filter will be d2 x d2r (where d2 is the size of the LBP difference patch)

        # LBP_Loop ---> IN
        ##########################################################################
        fileList = sorted(os.listdir(lbpPath))

        for L_i in range(self.LBP_list.shape[0]):
            if self.verbose:
                print('LBP ',str(L_i+1))

            # loop read all cases
            ######################################################################

            fileList_Normal = []
            fileList_Syndromic = []

            # Separating Normals and Syndromics
            for file in fileList:

                fileName, fileExtension = os.path.splitext(file)

                # if fileExtension == '.dat':

                if self.negativeClassName in fileName:
                    fileList_Normal.append(fileName)

                elif self.positiveClassName in fileName:
                    fileList_Syndromic.append(fileName)

            # Normals ------------------------------------------------------------
            for f_i in range(len(fileList_Normal)):

                #LBP_All_i = np.load(os.path.join(lbpPath,fileList_Normal[f_i]))
                shelveDictionary = shelve.open(os.path.join(lbpPath, fileList_Normal[f_i]))
                LBP_All_i = shelveDictionary['LBP_' + str(L_i+1)]
                shelveDictionary.close()

                LBP_All_i = np.reshape(LBP_All_i, (LBP_All_i.shape[0],LBP_All_i.shape[1],1,LBP_All_i.shape[2]))

                if f_i == 0:
                    LBP_ALL_Ldmks = np.copy(LBP_All_i)
                else:
                    LBP_ALL_Ldmks = np.append(LBP_ALL_Ldmks,LBP_All_i,axis=2)

            LBP_Normal = LBP_ALL_Ldmks

            # Syndromics ---------------------------------------------------------
            for f_i in range(len(fileList_Syndromic)):

                #LBP_All_i = np.load(os.path.join(lbpPath,fileList_Syndromic[f_i]))
                shelveDictionary = shelve.open(os.path.join(lbpPath, fileList_Syndromic[f_i]))
                LBP_All_i = shelveDictionary['LBP_' + str(L_i+1)]
                shelveDictionary.close()

                LBP_All_i = np.reshape(LBP_All_i, (LBP_All_i.shape[0],LBP_All_i.shape[1],1,LBP_All_i.shape[2]))

                if f_i == 0:
                    LBP_ALL_Ldmks = np.copy(LBP_All_i)
                else:
                    LBP_ALL_Ldmks = np.append(LBP_ALL_Ldmks,LBP_All_i,axis=2)

            LBP_Syndromic = LBP_ALL_Ldmks

            (d1,d2,C_norm,N_L) = LBP_Normal.shape
            C_synd = LBP_Syndromic.shape[2]

            # LEAVE ONE OUT LOOP ---> IN
            ######################################################################

            fileList_Names = fileList_Normal + fileList_Syndromic

            LBP_All = np.zeros((LBP_Normal.shape[0],LBP_Normal.shape[1],LBP_Normal.shape[2]+LBP_Syndromic.shape[2],LBP_Normal.shape[3]))

            LBP_All[:,:,:LBP_Normal.shape[2],:] = np.copy(LBP_Normal)
            LBP_All[:,:,LBP_Normal.shape[2]:,:] = np.copy(LBP_Syndromic)
            Normal_List_All = np.vstack((np.ones((LBP_Normal.shape[2],1)),np.zeros((LBP_Syndromic.shape[2],1))))
            Syndromic_List_All = np.vstack((np.zeros((LBP_Normal.shape[2],1)),np.ones((LBP_Syndromic.shape[2],1))))

            for sh_i in range(len(fileList_Names)):

                patientDictionary = {}

                if self.verbose:
                    print('Processing case: {}'.format(fileList_Names[sh_i]))
                    print('-------------------------------------------------')

                Normal_List_All_i = np.copy(Normal_List_All)
                Normal_List_All_i[sh_i,:] = 0
                Syndromic_List_All_i = np.copy(Syndromic_List_All)
                Syndromic_List_All_i[sh_i,:] = 0

                LBP_Normal = LBP_All[:,:,np.where(Normal_List_All_i.ravel()==1)[0],:]
                LBP_Syndromic = LBP_All[:,:,np.where(Syndromic_List_All_i.ravel()==1)[0],:]

                C_norm = int(np.sum(Normal_List_All_i))
                C_synd = int(np.sum(Syndromic_List_All_i))

                # Landmark LOOP ---> IN
                ##################################################################
                for l_i in range(N_L):

                    # LBP_Normal : (Img. filter patch) x  (LBP pattern) x (Num imgs) x (Num Lndks)
                    Dm_ip_normal = np.mean(LBP_Normal[:,:,:,l_i],axis=2)
                    Dm_ip_syndromic = np.mean(LBP_Syndromic[:,:,:,l_i],axis=2)
                    Dm_p = (np.sum(LBP_Normal[:,:,:,l_i],axis=2)+np.sum(LBP_Syndromic[:,:,:,l_i],axis=2))/(C_norm+C_synd)

                    # Iterations LOOP ---> IN
                    ##################################################################
                    W = np.eye(d1)
                    V = np.eye(d2)

                    for t_i in range(T):

                        # STEP 1
                        ##############################################################

                        # 4.1: Compute within class scatter matrices..................
                        Sw_1_normal = np.zeros((d1,d1), dtype=np.float32)
                        for c_i in range(C_norm):   # images loop
                            dI_ijp = LBP_Normal[:,:,c_i,l_i] - Dm_ip_normal
                            Sw_1_aux = np.dot(np.dot(np.dot(dI_ijp, V), V.T), dI_ijp.T)
                            Sw_1_normal = Sw_1_normal + Sw_1_aux

                        Sw_1_syndromic = np.zeros((d1,d1), dtype=np.float32)
                        for c_i in range(C_synd):   # images loop
                            dI_ijp = LBP_Syndromic[:,:,c_i,l_i] - Dm_ip_syndromic
                            Sw_1_aux = np.dot(np.dot(np.dot(dI_ijp, V), V.T), dI_ijp.T)
                            Sw_1_syndromic = Sw_1_syndromic + Sw_1_aux

                        Sw_1 = Sw_1_normal + Sw_1_syndromic

                        # 4.2: Compute between class scatter matrices...............
                        Sb_1 = (C_norm * np.dot(np.dot(np.dot(Dm_ip_normal-Dm_p, V), V.T), (Dm_ip_normal-Dm_p).T)) + (C_synd * np.dot(np.dot(np.dot(Dm_ip_syndromic-Dm_p, V), V.T), (Dm_ip_syndromic-Dm_p).T))

                        # 4.3: Solve the generalized eigenvalue problem...............
                        #   Sb_1*W = L * Sw_1*W

                        D_eig, V_eig = scipy.linalg.eig(Sb_1, Sw_1, left=False, right=True, overwrite_a=True, overwrite_b=True, check_finite=True)

                        D_sortIndices = np.argsort(D_eig)[::-1]
                        D_eig = D_eig[D_sortIndices]
                        V_eig = V_eig[D_sortIndices,:]

                        W = np.copy(V_eig[:,:d1r])

                        Tr_aux = np.trace(np.dot(np.linalg.inv(np.dot(np.dot(W.T, Sw_1), W)), np.dot(np.dot(W.T, Sb_1), W)))

                        if t_i == 0:
                            Trace_1 = np.array(Tr_aux)
                        else:
                            Trace_1 = np.append(Trace_1,Tr_aux)

                        # STEP 2:
                        ##############################################################
                        # 5.1: Compute the within class scatter matrix................
                        Sw_2_normal = np.zeros((d2,d2))
                        for c_i in range(C_norm):   # images loop
                            dI_ijp = LBP_Normal[:,:,c_i,l_i] - Dm_ip_normal
                            Sw_2_aux = np.dot(np.dot(np.dot(dI_ijp.T, W), W.T), dI_ijp)
                            Sw_2_normal = Sw_2_normal + Sw_2_aux

                        Sw_2_syndromic = np.zeros((d2,d2))
                        for c_i in range(C_synd):   # images loop
                            dI_ijp = LBP_Syndromic[:,:,c_i,l_i] - Dm_ip_syndromic
                            Sw_2_aux = np.dot(np.dot(np.dot(dI_ijp.T, W), W.T), dI_ijp)
                            Sw_2_syndromic = Sw_2_syndromic + Sw_2_aux

                        Sw_2 = Sw_2_normal + Sw_2_syndromic

                        # 5.2: Compute between class scatter matrices.................
                        Sb_2 = (C_norm * np.dot(np.dot(np.dot((Dm_ip_normal-Dm_p).T, W), W.T), (Dm_ip_normal-Dm_p))) + (C_synd * np.dot(np.dot(np.dot((Dm_ip_syndromic-Dm_p).T, W), W.T), (Dm_ip_syndromic-Dm_p)))
                        # print(Sb_2)
                        # 5.3: Solve the generalized eigenvalue problem...............
                        D_eig, V_eig = scipy.linalg.eig(Sb_2, Sw_2, left=False, right=True, overwrite_a=True, overwrite_b=True, check_finite=True)

                        D_sortIndices = np.argsort(D_eig)[::-1]
                        D_eig = D_eig[D_sortIndices]
                        V_eig = V_eig[D_sortIndices,:]


                        V = np.copy(V_eig[:,:d2r])

                        Tr_aux = np.trace(np.dot(np.linalg.inv(np.dot(np.dot(V.T, Sw_2), V)), np.dot(np.dot(V.T, Sb_2), V)))

                        if t_i == 0:
                            Trace_2 = np.array(Tr_aux)
                        else:
                            Trace_2 = np.append(Trace_2,Tr_aux)

                        ##################################################################
                        # Iterations LOOP ---> OUT

                    # Store the W and V for all the landmarks.
                    if l_i == 0:
                        W_list = np.zeros((W.shape[0],W.shape[1],1,N_L))
                        V_list = np.zeros((V.shape[0],V.shape[1],1,N_L))

                    W_list[:,:,0,l_i] = np.copy(W).real
                    V_list[:,:,0,l_i] = np.copy(V).real

                    ##############################################################
                    # Landmkark LOOP ---> OUT

                shelveDictionary = shelve.open(os.path.join(outputPath, fileList_Names[sh_i]))
                shelveDictionary['W_' + str(L_i + 1)] = W_list
                shelveDictionary['V_' + str(L_i + 1)] = V_list
                shelveDictionary.close()


    # Returns the geometric features from a list of landmarks
    def calculateGeometricFeatures(self, landmarkList):
        # landmarkList: list of landmarks with shape [nSubjects, nLandmarks, 2]

        nImages = landmarkList.shape[0]

        geometricFeatures = np.ndarray([nImages, 0], dtype=np.float32)
        featureDescription = []

        # Normalization factors
        hBaseline = np.sqrt((landmarkList[:,7,0]-landmarkList[:,0,0])**2 + (landmarkList[:,7,1]-landmarkList[:,0,1])**2) # Distance between lateral canthi
        vBaseline = np.sqrt((landmarkList[:,28,0]-landmarkList[:,7,0])**2 + (landmarkList[:,28,1]-landmarkList[:,7,1])**2) # Distance between lateral canthi and oral commissures

        # Horizontal features
        rightDist = np.sqrt( (landmarkList[:,2,0] - landmarkList[:,0,0])**2 + (landmarkList[:,2,1] - landmarkList[:,0,1])**2) / hBaseline
        leftDist = np.sqrt( (landmarkList[:,7,0] - landmarkList[:,5,0])**2 + (landmarkList[:,7,1] - landmarkList[:,5,1])**2) / hBaseline
        average = (rightDist + leftDist) / 2.0

        if self.asymmetry:
            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in distance between medial and lateral canthi')

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        featureDescription.append('Distance between medial and lateral canthi')

        rightDist = np.sqrt( (landmarkList[:,5,0] - landmarkList[:,2,0])**2 + (landmarkList[:,5,1] - landmarkList[:,2,1])**2)  / hBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between medial canthi')

        rightDist = np.sqrt( (landmarkList[:,18,0] - landmarkList[:,12,0])**2 + (landmarkList[:,18,1] - landmarkList[:,12,1])**2) / hBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between nose alae')

        rightDist = np.sqrt( (landmarkList[:,28,0] - landmarkList[:,24,0])**2 + (landmarkList[:,28,1] - landmarkList[:,24,1])**2) / hBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between oral commissures')

        # Vertical features
        rightDist = np.sqrt( (landmarkList[:,2,0] - landmarkList[:,15,0])**2 + (landmarkList[:,2,1] - landmarkList[:,15,1])**2) / vBaseline
        leftDist = np.sqrt( (landmarkList[:,5,0] - landmarkList[:,15,0])**2 + (landmarkList[:,5,1] - landmarkList[:,15,1])**2) / vBaseline
        average = (rightDist + leftDist) / 2.0
        #asymmetry = np.sqrt((rightDist - average)**2 + (leftDist - average)**2)

        if self.asymmetry:

            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in distance between medial canthi and columella')

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        featureDescription.append('Distance between medial canthi and columella')

        rightDist = np.sqrt( (landmarkList[:,21,0] - landmarkList[:,15,0])**2 + (landmarkList[:,21,1] - landmarkList[:,15,1])**2) / vBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Nose length')

        rightDist = np.sqrt( (landmarkList[:,15,0] - landmarkList[:,26,0])**2 + (landmarkList[:,15,1] - landmarkList[:,26,1])**2) / vBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between columella and cupid\'s bow')

        rightDist = np.sqrt( (landmarkList[:,31,0] - landmarkList[:,25,0])**2 + (landmarkList[:,31,1] - landmarkList[:,25,1])**2) / vBaseline
        leftDist = np.sqrt( (landmarkList[:,31,0] - landmarkList[:,27,0])**2 + (landmarkList[:,31,1] - landmarkList[:,27,1])**2) / vBaseline
        average = (rightDist + leftDist) / 2.0
        #asymmetry = np.sqrt((rightDist - average)**2 + (leftDist - average)**2)

        if self.asymmetry:
            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in upper lip width')

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        featureDescription.append('Upper lip width')

        rightDist = np.sqrt( (landmarkList[:,30,0] - landmarkList[:,32,0])**2 + (landmarkList[:,30,1] - landmarkList[:,32,1])**2) / vBaseline
        leftDist = np.sqrt( (landmarkList[:,29,0] - landmarkList[:,32,0])**2 + (landmarkList[:,29,1] - landmarkList[:,32,1])**2) / vBaseline
        average = (rightDist + leftDist) / 2.0
        #asymmetry = np.sqrt((rightDist - average)**2 + (leftDist - average)**2)

        if self.asymmetry:

            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in lower lip width')

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        featureDescription.append('Lower lip width')

        #New Features for Sickle Cell
        midPoint_x = (landmarkList[:,30,0] + landmarkList[:,29,0])/(2 * hBaseline)
        midPoint_y =  (landmarkList[:,30,1] +  landmarkList[:,29,1])/(2 * hBaseline)

        rightDist = np.sqrt((midPoint_x - landmarkList[:,15,0])**2 + (midPoint_y - landmarkList[:,15,1])**2) / vBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between Columella and Lower lip')

        rightDist = np.sqrt((landmarkList[:,21,0] - landmarkList[:,23,0])**2 + (landmarkList[:,21,1] - landmarkList[:,23,1])**2) / vBaseline
        geometricFeatures = np.append(geometricFeatures, rightDist.reshape([rightDist.size,1]), axis=1)
        featureDescription.append('Distance between nasion and philtrum')

        # Angles
        points = [0, 2, 5]
        x1 = landmarkList[:,points[0],0]-landmarkList[:,points[1],0]
        x2 = landmarkList[:,points[1],0]-landmarkList[:,points[2],0]
        y1 = landmarkList[:,points[0],1]-landmarkList[:,points[1],1]
        y2 = landmarkList[:,points[1],1]-landmarkList[:,points[2],1]
        v1 = np.array([x1,y1], dtype=np.float32) / np.sqrt(x1**2 + y1**2) # vector left medial canthus -> left lateral canthus
        v2 = np.array([x2,y2], dtype=np.float32) / np.sqrt(x2**2 + y2**2) # vector right medial canthus -> left medial canthus
        righDist = (np.arcsin(v1[1]) - np.arcsin(v2[1]))*180/np.pi

        points = [2, 5, 7]
        x1 = landmarkList[:,points[1],0]-landmarkList[:,points[0],0]
        x2 = landmarkList[:,points[2],0]-landmarkList[:,points[1],0]
        y1 = landmarkList[:,points[1],1]-landmarkList[:,points[0],1]
        y2 = landmarkList[:,points[2],1]-landmarkList[:,points[1],1]
        v1 = np.array([x1,y1], dtype=np.float32) / np.sqrt(x1**2 + y1**2) # vector left medial canthus -> right medial canthus
        v2 = np.array([x2,y2], dtype=np.float32) / np.sqrt(x2**2 + y2**2) # vector right medial canthus -> right lateral canthus
        leftDist = (np.arcsin(v2[1]) - np.arcsin(v1[1]))*180/np.pi

        average = (rightDist + leftDist) / 2.0
        #asymmetry = np.sqrt((rightDist - average)**2 + (leftDist - average)**2)
        #asymmetry = np.abs(rightDist - leftDist)

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        #geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size,1]), axis=1)
        featureDescription.append('Eye slanting')

        if self.asymmetry:
            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in eye slanting')

        points = [13, 12, 11]
        x1 = landmarkList[:,points[0],0]-landmarkList[:,points[1],0]
        x2 = landmarkList[:,points[2],0]-landmarkList[:,points[1],0]
        y1 = landmarkList[:,points[0],1]-landmarkList[:,points[1],1]
        y2 = landmarkList[:,points[2],1]-landmarkList[:,points[1],1]
        righDist = (x1*x2 + y1*y2)/( (np.sqrt(x1**2 + y1**2))* (np.sqrt(x2**2 + y2**2)) + np.spacing(1))
        if np.abs(rightDist) > 1: # This corrects rounding inaccuracies
            rightDist = np.sign(rightDist)*1
        righDist = np.arccos(rightDist)*180/np.pi
        points = [17, 18, 19]
        x1 = landmarkList[:,points[0],0]-landmarkList[:,points[1],0]
        x2 = landmarkList[:,points[2],0]-landmarkList[:,points[1],0]
        y1 = landmarkList[:,points[0],1]-landmarkList[:,points[1],1]
        y2 = landmarkList[:,points[2],1]-landmarkList[:,points[1],1]
        leftDist = (x1*x2 + y1*y2)/( (np.sqrt(x1**2 + y1**2))* (np.sqrt(x2**2 + y2**2)) + np.spacing(1))
        if np.abs(leftDist) > 1: # This corrects rounding inaccuracies
            leftDist = np.sign(leftDist)*1
        leftDist = np.arccos(leftDist)*180/np.pi
        average = (rightDist + leftDist) / 2.0
        #asymmetry = np.sqrt((rightDist - average)**2 + (leftDist - average)**2)
        if self.asymmetry:

            asymmetry = np.abs(rightDist - leftDist)
            geometricFeatures = np.append(geometricFeatures, asymmetry.reshape([asymmetry.size, 1]), axis=1)
            featureDescription.append('Asymmetry in the angle at the alas of the nose')

        geometricFeatures = np.append(geometricFeatures, average.reshape([average.size,1]), axis=1)
        featureDescription.append('Angle at the alas of the nose')

        return geometricFeatures, featureDescription

    # Returns the features for a list of subjects and their description
    def GetFeatures(self, imageList, landmarks, LDAMatrices):
        # imageList is a list of images
        # landmarks is an array with shape len(images) x nLandmarks x 2
        # LDAMatrices is a dictionary with keys W_1, V_1, W_2, V_2, W_3, V_3, where 1,2,3 are the three resolutions of the LBPs

        geometricFeatures, featureDescriptions = self.calculateGeometricFeatures(landmarks)
        # geometricFeatures is an array with shape nPatients x nGeometricFeatures

        for i in range(len(imageList)):

            lbpDictionary = self.extractLBP(imageList[i], landmarks[i,:,:], self.LBP_list, self.Ws)

            texture_features = []


            for value in lbpDictionary["FeatureDescriptions"]:
                texture_features.append(value)
                # if "Asymmetry" not in value:
                #     texture_features.append(value)

            lbpDictionary["FeatureDescriptions"] = texture_features

            patientTextureFeatures = np.ndarray([0], dtype=np.float32)
            for LBP_resolution in range(self.LBP_list.shape[0]):

                lbpShape = lbpDictionary['LBP_'+str(LBP_resolution+1)].shape
                vShape = LDAMatrices['V_'+str(LBP_resolution+1)].shape

                LBP_mat = np.reshape(lbpDictionary['LBP_'+str(LBP_resolution+1)], (lbpShape[0], lbpShape[1], 1,lbpShape[2]))

                W_mat = np.tile(LDAMatrices['W_'+str(LBP_resolution+1)], [1, lbpShape[1], 1, 1])

                V_mat = np.reshape(LDAMatrices['V_'+str(LBP_resolution+1)], (vShape[1], vShape[0], vShape[2], vShape[3]))

                patientTextureFeatures = np.concatenate([patientTextureFeatures, np.sum(np.sum(W_mat*LBP_mat, axis=0) * V_mat, axis=1).ravel()])

                # Adding the description of features
                if i==0:
                    for j in range(len(lbpDictionary['FeatureDescriptions'])):
                        featureDescriptions += [lbpDictionary['FeatureDescriptions'][j] + ' (R' + str(LBP_resolution+1) + ')']

            patientFeatures = np.concatenate([geometricFeatures[i,:], patientTextureFeatures]).reshape(1, patientTextureFeatures.size+geometricFeatures.shape[1])

            if i==0:
                allFeatures = patientFeatures
            else:
                allFeatures = np.concatenate([allFeatures, patientFeatures], axis=0)

        return allFeatures, featureDescriptions

    # Plots the texture at the indicated landmark
    def GetTextureAtLandmark(self, workingPath, landmarkId):

        centerIndices = np.array([15, 21, 22, 23, 26, 31, 32], dtype=np.int)
        leftIndices  = np.array([0, 1, 2, 3, 4, 10, 11, 12, 13, 14, 24, 25, 30], dtype=np.int)
        rightIndices = np.array([7, 6, 5, 8, 9, 20, 19, 18, 17, 16, 28, 27, 29], dtype=np.int)

        listOfImagesToReturn = [] # List of all images returned.
        # Format: negativeClassAverageR1, negativeClassAverageR2, negativeClassAverageR3.
        # positiveClassAverageR1, positiveClassAverageR2, positiveClassAverageR3,
        # If landmarkId has a symmetric landmarks, the following is appended
        # negativeClassDifferenceR1, negativeClassDifferenceR2, negativeClassDifferenceR3.
        # positiveClassDifferenceR1, positiveClassDifferenceR2, positiveClassDifferenceR3,

        symmetricId = None
        if landmarkId in leftIndices:
            symmetricId = rightIndices[np.argwhere(leftIndices==landmarkId)[0]][0]
        elif landmarkId in rightIndices:
            tmp = landmarkId
            symmetricId = landmarkId
            # landmarkId = rightIndices[nps.argwhere(rightIndices==symmetricId)[0]][0]

        # List of images, one per resolution, for the negative class
        negativeAverageImages = []
        negativeDifferenceImages = []
        numberOfNegativeCases = 0

        negativeMatrixData = []
        positiveMatrixData = []

        # Reading the negative class images
        imageFolder = os.path.join(workingPath, self.resizedImageFolder, self.negativeClassName)
        landmarksFolder = os.path.join(imageFolder, self.landmarksFolder)
        fileList = os.listdir(imageFolder)
        for file in fileList:

            fileName, fileExtension = os.path.splitext(file)

            if fileExtension.lower() == '.jpg':

                # Read image
                image = plt.imread(os.path.join(imageFolder, file),'L')
                # Normalizing image values between 0 and 1 --> equivalent to MATLAB mat2gray function
                min_ = np.min(image)
                max_ = np.max(image)
                image = (image - min_) / (max_ - min_ + np.spacing(1))

                # Read landmarks
                landmarks = np.load(os.path.join(landmarksFolder, fileName+'.npy'))

                # Expand the image to avoid accessing to the wrong location in case
                # the filter on the LBP patch were too big.
                P_Extra = np.max(np.array([self.LBP_list[0,1],self.LBP_list[1,1],self.LBP_list[2,1],self.Ws])) + 5

                I_Corner_UL = image[(P_Extra-1)::-1,(P_Extra-1)::-1]
                I_Corner_UR = image[(P_Extra-1)::-1,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]
                I_Corner_BL = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,(P_Extra-1)::-1]
                I_Corner_BR = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]
                I_u = image[(P_Extra-1)::-1,:]
                I_b = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,:]
                I_l = image[:,(P_Extra-1)::-1]
                I_r = image[:,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]

                image = np.vstack((np.hstack((I_Corner_UL,I_u,I_Corner_UR)),np.hstack((I_l,image,I_r)),np.hstack((I_Corner_BL,I_b,I_Corner_BR))))

                landmark = np.round(landmarks[landmarkId, :] + P_Extra).astype(np.int) # This is for left and center indices

                if symmetricId is not None:
                    symmetricLandmark = np.round(landmarks[symmetricId, :] + P_Extra).astype(np.int) # This is for left and center indices

                # For each resolution level
                for resolution in range(self.LBP_list.shape[0]):

                    R = np.max([int(self.LBP_list[resolution, 1] + (self.Ws-1)/2), 24])

                    # Left (or center)
                    leftPatch = image[landmark[1]-R:landmark[1]+R+1, landmark[0]-R:landmark[0]+R+1]
                    leftPatch = (leftPatch-np.min(leftPatch))/(np.max(leftPatch)-np.min(leftPatch))

                    if numberOfNegativeCases == 0:
                        negativeAverageImages.append(np.copy(leftPatch))
                    else:
                        negativeAverageImages[resolution] += np.copy(leftPatch)

                    if symmetricId is not None:
                        rightPatch = np.fliplr(image[symmetricLandmark[1]-R:symmetricLandmark[1]+R+1, symmetricLandmark[0]-R:symmetricLandmark[0]+R+1])
                        rightPatch = (rightPatch-np.min(rightPatch))/(np.max(rightPatch)-np.min(rightPatch))
                        negativeAverageImages[resolution] += np.copy(rightPatch)

                        if numberOfNegativeCases == 0:
                            negativeDifferenceImages.append(np.abs(leftPatch - rightPatch))
                        else:
                            negativeDifferenceImages[resolution] += np.abs(leftPatch - rightPatch)

                    if numberOfNegativeCases == 0:
                        if symmetricId is not None:
                            negativeMatrixData.append(((leftPatch + rightPatch)/2).reshape(1,leftPatch.size))
                        else:
                            negativeMatrixData.append(leftPatch.reshape(1,leftPatch.size))
                    else:
                        if symmetricId is not None:
                            negativeMatrixData[resolution] = np.concatenate([negativeMatrixData[resolution], ((leftPatch + rightPatch)/2).reshape(1,leftPatch.size)], axis=0)
                        else:
                            negativeMatrixData[resolution] = np.concatenate([negativeMatrixData[resolution], leftPatch.reshape(1,leftPatch.size)], axis=0)

                numberOfNegativeCases += 1

        for resolution in range(self.LBP_list.shape[0]):
            negativeAverageImages[resolution] /= numberOfNegativeCases
            if symmetricId is not None:
                negativeDifferenceImages[resolution] /= numberOfNegativeCases
                negativeAverageImages[resolution] /= 2

        ##################################################

        # List of images, one per resolution, for the positve class
        positiveAverageImages = []
        positiveDifferenceImages = []
        numberOfPositiveCases = 0

        # Reading the positive class images
        imageFolder = os.path.join(workingPath, self.resizedImageFolder, self.positiveClassName)
        landmarksFolder = os.path.join(imageFolder, self.landmarksFolder)
        fileList = os.listdir(imageFolder)
        for file in fileList:

            fileName, fileExtension = os.path.splitext(file)
            # print(file)

            if fileExtension.lower() == '.jpg':

                # Read image
                image = plt.imread(os.path.join(imageFolder, file),'L')
                # Normalizing image values between 0 and 1 --> equivalent to MATLAB mat2gray function
                min_ = np.min(image)
                max_ = np.max(image)
                image = (image - min_) / (max_ - min_ + np.spacing(1))

                # Read landmarks
                landmarks = np.load(os.path.join(landmarksFolder, fileName+'.npy'))

                # Expand the image to avoid accessing to the wrong location in case
                # the filter on the LBP patch were too big.
                P_Extra = np.max(np.array([self.LBP_list[0,1],self.LBP_list[1,1],self.LBP_list[2,1],self.Ws])) + 5
                # print(P_Extra)

                I_Corner_UL = image[(P_Extra-1)::-1,(P_Extra-1)::-1]
                I_Corner_UR = image[(P_Extra-1)::-1,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]
                I_Corner_BL = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,(P_Extra-1)::-1]
                I_Corner_BR = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]
                I_u = image[(P_Extra-1)::-1,:]
                I_b = image[(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1,:]
                I_l = image[:,(P_Extra-1)::-1]
                I_r = image[:,(image.shape[1]-1):(image.shape[1]-1)-(P_Extra+2):-1]

                image = np.vstack((np.hstack((I_Corner_UL,I_u,I_Corner_UR)),np.hstack((I_l,image,I_r)),np.hstack((I_Corner_BL,I_b,I_Corner_BR))))

                landmark = np.round(landmarks[landmarkId, :] + P_Extra).astype(np.int) # This is for left and center indices
                # landmark =  np.round(landmarks[landmarkId, :]).astype(np.int)

                if symmetricId is not None:
                    symmetricLandmark = np.round(landmarks[symmetricId, :] + P_Extra).astype(np.int) # This is for left and center indices
                    # symmetricLandmark = np.round(landmarks[symmetricId, :]).astype(np.int)

                # For each resolution level
                for resolution in range(self.LBP_list.shape[0]):

                    R = np.max([int(self.LBP_list[resolution, 1] + (self.Ws-1)/2), 24])
                    # print(R)

                    # Left (or center)
                    leftPatch = image[landmark[1]-R:landmark[1]+R+1, landmark[0]-R:landmark[0]+R+1]
                    leftPatch = (leftPatch-np.min(leftPatch))/(np.max(leftPatch)-np.min(leftPatch))
                    if numberOfPositiveCases == 0:
                        positiveAverageImages.append(np.copy(leftPatch))
                    else:
                        positiveAverageImages[resolution] += np.copy(leftPatch)

                    if symmetricId is not None:
                        rightPatch = np.fliplr(image[symmetricLandmark[1]-R:symmetricLandmark[1]+R+1, symmetricLandmark[0]-R:symmetricLandmark[0]+R+1])
                        rightPatch = (rightPatch-np.min(rightPatch))/(np.max(rightPatch)-np.min(rightPatch))
                        positiveAverageImages[resolution] += np.copy(rightPatch)

                        if numberOfPositiveCases == 0:
                            positiveDifferenceImages.append(np.abs(leftPatch - rightPatch))
                        else:
                            positiveDifferenceImages[resolution] += np.abs(leftPatch - rightPatch)

                    if numberOfPositiveCases == 0:
                        if symmetricId is not None:
                            positiveMatrixData.append(((leftPatch + rightPatch)/2).reshape(1,leftPatch.size))
                        else:
                            positiveMatrixData.append(leftPatch.reshape(1,leftPatch.size))
                    else:
                        if symmetricId is not None:
                            positiveMatrixData[resolution] = np.concatenate([positiveMatrixData[resolution], ((leftPatch + rightPatch)/2).reshape(1,leftPatch.size)], axis=0)
                        else:
                            positiveMatrixData[resolution] = np.concatenate([positiveMatrixData[resolution], leftPatch.reshape(1,leftPatch.size)], axis=0)


                numberOfPositiveCases += 1

        for resolution in range(self.LBP_list.shape[0]):
            positiveAverageImages[resolution] /= numberOfPositiveCases
            if symmetricId is not None:
                positiveDifferenceImages[resolution] /= numberOfPositiveCases
                positiveAverageImages[resolution] /= 2

        #####
        listOfImagesToReturn += negativeAverageImages
        listOfImagesToReturn += positiveAverageImages

        if symmetricId is not None:
            listOfImagesToReturn += negativeDifferenceImages
            listOfImagesToReturn += positiveDifferenceImages

        return listOfImagesToReturn


# Auxiliary function to calculate the local binary patterns
def generateRadialFilterLBP(p=8, r=1):
    # Rewrote MATLAB SCRIPT
#    r = max((1,r) # radius below 1 is illegal
#    p = np.round(p) # non integer number of neighbors
#    p = max(1,p)

    # Find elements angles, arranged counter-clockwise starting from "x-axis"
    theta = np.linspace(0,2*np.pi,p+1)+np.pi/2
    theta = theta[:(theta.shape[0]-1)].reshape(1,theta.shape[0]-1) # Removes last element (0=2pi)

    ## Find relevant coordinates
    rowsFilt,colsFilt = pol2cart(np.matlib.repmat(r,theta.shape[0],theta.shape[1]),theta)
    nEps = -3
    rowsFilt = np.round(rowsFilt/(10**nEps)) * (10**nEps)
    colsFilt = np.round(colsFilt/(10**nEps)) * (10**nEps)

    # Matrix indices should be integers
    rowsFloor = np.floor(rowsFilt)
    rowsCeil = np.ceil(rowsFilt)
    colsFloor = np.floor(colsFilt)
    colsCeil = np.ceil(colsFilt)

    rowsDistFloor = 1-np.abs(rowsFloor-rowsFilt)
    rowsDistCeil = 1-np.abs(rowsCeil-rowsFilt)
    colsDistFloor = 1-np.abs(colsFloor-colsFilt)
    colsDistCeil = 1-np.abs(colsCeil-colsFilt)

    # Find minimal filter dimensions, basec on indices
    filtDims = np.array([(np.ceil(np.max(rowsFilt[0,:]))-np.floor(np.min(rowsFilt[0,:]))),(np.ceil(np.max(colsFilt[0,:]))-np.floor(np.min(colsFilt[0,:])))],dtype=np.float32).reshape(1,2)
    filtDims = filtDims + np.mod(filtDims+1,2)
    filtCenter = (filtDims+1)/2

    ## Convert cartesian coordinates to matrix elements coordinates via simple shift
    rowsFloor = rowsFloor+filtCenter[0,0]
    rowsCeil = rowsCeil+filtCenter[0,0]
    colsFloor = colsFloor+filtCenter[0,1]
    colsCeil = colsCeil+filtCenter[0,1]

    ## Generate the filter - each 2D slice for filter element
    radInterpFilt = np.zeros((int(filtDims[0,0]),int(filtDims[0,1]),p))
    for iP in range(p):
        radInterpFilt[int(rowsFloor[0,iP])-1,int(colsFloor[0,iP])-1,iP] = radInterpFilt[int(rowsFloor[0,iP])-1,int(colsFloor[0,iP])-1,iP] + rowsDistFloor[0,iP] + colsDistFloor[0,iP]

        radInterpFilt[int(rowsFloor[0,iP])-1,int(colsCeil[0,iP])-1,iP] = radInterpFilt[int(rowsFloor[0,iP])-1,int(colsCeil[0,iP])-1,iP] + rowsDistFloor[0,iP] + colsDistCeil[0,iP]

        radInterpFilt[int(rowsCeil[0,iP])-1,int(colsFloor[0,iP])-1,iP] = radInterpFilt[int(rowsCeil[0,iP])-1,int(colsFloor[0,iP])-1,iP] + rowsDistCeil[0,iP] + colsDistFloor[0,iP]

        radInterpFilt[int(rowsCeil[0,iP])-1,int(colsCeil[0,iP])-1,iP] = radInterpFilt[int(rowsCeil[0,iP])-1,int(colsCeil[0,iP])-1,iP] + rowsDistCeil[0,iP] + colsDistCeil[0,iP]

        radInterpFilt[:,:,iP] = radInterpFilt[:,:,iP]/(np.sum(radInterpFilt[:,:,iP]))

    radInterpFilt[int(filtCenter[0,0])-1,int(filtCenter[0,1])-1,:] = radInterpFilt[int(filtCenter[0,0])-1,int(filtCenter[0,1])-1,:] - 1

    return radInterpFilt

# Converts from cartesian to polar coordinates
def cart2pol(x, y):
    # Returns angle, magnitude
    return np.float32(np.arctan2(y, x)), np.float32(np.sqrt(x**2 + y**2))

# Converts from polar to cartesian coordinates
def pol2cart(rho, phi):
    # Returns x, y
    return rho * np.cos(phi), rho * np.sin(phi)

# Features selector
def MCFS_p(fea, FeaNumCandi, options):
    # FeaIndex = MCFS_p(scaledData, k, options)
    # MCFS ---> Feature Selection for Multi Class/Cluster data (rewrote MATLAB script)
    # Input:
    #   fea --> Data matrix (each row vector of diata is a sample vector)
    #   FeaNumCandi --> Number of features to be selected
    #   Options --> options_gnd (label of the data, used if a supervised feature selection problem)
    #
    # Output:
    #   FeaIndex --> Each element in FEAIndex is the index of the selected features
    (nSmp, nFea) = fea.shape
    FeaNumCandi = np.unique(FeaNumCandi)
    FeaNumCandi = np.copy(FeaNumCandi)

    nUseEigenfunction = 5.0
    if 'nUseEigenfunction' in options.keys():
        nUseEigenfunction = options['nUseEigenfunction']

    k = 5.0
    if 'k' in options.keys():
        k = options['k']

    if 'ratio' in options.keys():
        ratio = options['ratio']
    else:
        ratio = 1.0

    if 'NotEnoughNonZero' in options.keys():
        NotEnoughNonZero = options['NotEnoughNonZero']
    else:
        NotEnoughNonZero = 3.0

    if 'gnd' in options.keys():
        assert (options['gnd'].size == nSmp),"gnd does not match"
        gnd = np.copy(options['gnd'])
        classLabel = np.unique(gnd)
        classLabel = np.copy(classLabel).reshape(classLabel.shape[0],1)
        nClass = classLabel.shape[0]

        np.random.seed(0)
        Y = np.random.rand(nClass,nClass)
        Z = np.zeros((nSmp,nClass))
        for i in range(nClass):
            idx = np.argwhere(gnd[:] == classLabel[i])
            Z[idx[:,0],:] = np.matlib.repmat(Y[i,:],idx.shape[0],1)
        Z[:,0] = np.ones(nSmp)
        Y_,R = scipy.linalg.qr(Z, overwrite_a=False, lwork=None, mode='economic', pivoting=False, check_finite=True)
        Y = Y_[:,1].reshape(Y_.shape[0],1)

    options['ReguType'] = 'RidgeLasso'
    if 'Method' not in options.keys():
        options['Method'] = 'LASSO_LARs'

    if options['Method'].lower() == 'LASSO_LARs'.lower():
        options['LASSOway'] = 'LARs'
        options['LassoCardi'] = np.ceil((FeaNumCandi[0])*ratio).reshape(1,FeaNumCandi.shape[0])
        eigvectorAll = SR(options, Y, fea)

        FeaIndex = {}
        for i in range(FeaNumCandi.size):
            eigvector = np.copy(eigvectorAll[i])
            eigvector = np.amax(np.abs(eigvector),axis=1).reshape(eigvector.shape[0],1)

            idx = np.argsort(eigvector[:,0])[::-1].reshape(eigvector.shape)
            dump = np.zeros(idx.shape)
            for ii in range(idx.shape[0]):
                dump[ii,:] = eigvector[idx[ii],:]
            if dump[int(FeaNumCandi[i]-1), 0] == 0:
                assert (NotEnoughNonZero != 0),"Not Enough Fea!"
                assert (NotEnoughNonZero != 1),"Not Enough FEA!"

                for j in range(i,FeaNumCandi.size):
                    eigvec = np.copy(eigvectorAll[j])
                    eigvec = np.amax(np.abs(eigvec),axis=1).reshape(eigvec.shape[0],1)

                    idx2 = np.argsort(eigvec[:,0])[::-1]
                    dump2 = np.zeros(idx2.shape)
                    for jj in range(idx2.shape[0]):
                        dump2[jj] = eigvec[idx2[jj],:][0]
                    if dump2[int(FeaNumCandi[i])] > 0:
                        break

                if dump2[int(FeaNumCandi[i])-1] > 0:
                    idx = np.copy(idx2)
                else:
                    assert (NotEnoughNonZero == 2),"Not enough FEA, tried to find more but failed."
                    idx = np.copy(idx2)

            FeaIndex[i] = idx[:FeaNumCandi[i]][:]

        return FeaIndex

# Auxiliary function of MCFS_p feature selector
def SR(options, responses, data):
    # eigvectorAll = SR(options, Y, fea)
    # SR ---> Spectral Regression
    # Input:
    #   data --> Data matrix (each row vector of data is a sample vector)
    #   Responses --> Response vectors (each column is a response vector)
    #   options
    # Output:
    #   Eigenvector
    #   Lassocardi

    MAX_MATRIX_SIZE = 10000 # Can change according to machine computational power
    if 'MAX_MATRIX_SIZE' in options.keys():
        MAX_MATRIX_SIZE = options['MAX_MATRIX_SIZE']

    if 'ReguType' not in options.keys():
        options['ReguType'] = 'Ridge'

    (nSmp, nFea) = data.shape

    if options['ReguType'].lower() == 'RidgeLasso'.lower():
        if 'ReguAlpha' not in options.keys():
            options['ReguAlpha'] = 0.05
        if 'RidgeAlpha' not in options.keys():
            options['RidgeAlpha'] = 0.001
        if 'LASSOway' not in options.keys():
            options['LASSOway'] = 'SLEP'
        if options['LASSOway'].lower() == 'LARs'.lower():
            if 'LassoCardi' in options.keys():
                LassoCardi = np.copy(options['LassoCardi'])
            else:
                LassoCardi = np.linspace(10,50,5).reshape(1,LassoCardi.shape[0])

    if options['ReguType'].lower() == 'RidgeLasso'.lower():
        nVector = responses.shape[1]
        if options['LASSOway'].lower() == 'LARs'.lower():
            eigvector = {}

            if nFea < MAX_MATRIX_SIZE:
                gram = np.dot(data.T,data)
                gram = np.maximum(gram,gram.T)

                if options['RidgeAlpha'] > 0:
                    for i in range(gram.shape[0]):
                        gram[i,i] = gram[i,i] + options['RidgeAlpha']


                for i in range(nVector):
                    model = sklearn.linear_model.Lars(fit_intercept=True,verbose=False,normalize=True,precompute=gram,n_nonzero_coefs=int(LassoCardi[0]),eps=np.finfo(np.float).eps,copy_X=True,fit_path=False)
                    model.fit(data,responses[:,i])
                    eigvector_full = model.coef_.T
                    eigvector_2sp =  scipy.sparse.csr_matrix(eigvector_full)
                    numC = eigvector_2sp.nnz

                    eigvector[i] = eigvector_2sp.copy()
            else:
                if options['RidgeAlpha'] > 0:
                    data = scipy.sparse.vstack([np.copy(data), np.sqrt(options['RidgeAlpha'])*scipy.sparse.eye(nFea)])
                    responses = np.vstack((responses, np.zeros((nFea,nVector))))


                for i in range(nVector):
                    model = sklearn.linear_model.Lars(fit_intercept=True,verbose=False,normalize=True,precompute=gram,n_nonzero_coefs=int(LassoCardi[0]),eps=np.finfo(np.float).eps,copy_X=True,fit_path=False)
                    model.fit(data,responses[:,i])
                    eigvector_full = model.coef_.T
                    eigvector_2sp =  scipy.sparse.csr_matrix(eigvector_full)
                    numC = eigvector_2sp.nnz

                    eigvector[i] = eigvector_2sp.copy()

    if (options['ReguType'].lower() == 'RidgeLasso'.lower()) and (options['LASSOway'].lower() == 'LARs'.lower()):
        eigvectorAll = eigvector.copy()
        eigvector = {}

        for i in range(len(eigvectorAll)):
            eigvector_T = scipy.sparse.csr_matrix.toarray(eigvectorAll[i])
            (tm,tn) = eigvector_T.shape
            tCar = np.zeros((tn,1))
            for k in range(tn):
                ind = np.where(eigvector_T[:,k] != 0)
                tCar[k,0] = ind[0].size

            for cardidx in range(LassoCardi.size):
                ratio = LassoCardi[cardidx,0]
                iMin = np.where(tCar == ratio)
                if (iMin[0].size > 0) and (iMin[1].size > 0):
                    tmpEigvec = (eigvector_T[:,int(iMin[1][-1])]/(np.linalg.norm(eigvector_T[:,int(iMin[1][-1])]))).reshape(eigvector_T.shape)
                    eigvector[cardidx] = np.copy(tmpEigvec)

    return eigvector

def resizeImageAndLandmarks(inputImage, inputLandmarks, cnnImageSize, totalPadding=0):

    image = np.array(inputImage)
    coords = np.array(inputLandmarks)
    #coords = np.array(inputLandmarks[:,0:33,:])

    originalImageSize = (image.shape[0], image.shape[1])

    landmarksBoxCenter = (
        ( np.max(coords[0,:,0]) + np.min(coords[0,:,0]) ) / 2.0,
        ( np.max(coords[0,:,1]) + np.min(coords[0,:,1]) ) / 2.0 - (np.max(coords[0,:,1]) - np.min(coords[0,:,1]))*0.1 # The second term is a correction to center all
    )


    #######################################################################
    # This is used when we only want to extract the face
    landmarksBoxSize = (
        (np.max(coords[0,:,0]) - np.min(coords[0,:,0])),
        (np.max(coords[0,:,1]) - np.min(coords[0,:,1]))
    )

    tmp = np.max([landmarksBoxSize[0], landmarksBoxSize[1]])

    tmp += totalPadding

    if tmp%2 == 0: # Making the number odd
        tmp = tmp - 1
    landmarksBoxSize = (tmp, tmp)

    # This is used when we want the largest image
    ##tmp = np.min([originalImageSize[0], originalImageSize[1]])

    ##if tmp%2 == 0: # Making the number odd
    ##    tmp = tmp - 1

    ##landmarksBoxSize = (tmp, tmp)
    ########################################################################

    ## Rearranging the image

    # if we have to add zeros on the left
    if landmarksBoxCenter[0]-int(landmarksBoxSize[0]/2) < 0:
        tmp = int(np.ceil( int(landmarksBoxSize[0]/2) - landmarksBoxCenter[0]))

        if image.ndim == 2:
            padWidth = ( (0,0),(tmp, 0) )
        elif image.ndim == 3:
            padWidth = ( (0,0), (tmp, 0), (0, 0) )

        image = np.pad(image, pad_width=padWidth, mode='edge' )
        landmarksBoxCenter = (landmarksBoxCenter[0] + tmp, landmarksBoxCenter[1])
        for l in range(0, 44):
            coords[0][l][0] = coords[0][l][0] + tmp

    # If we have to add zeros on the right
    if landmarksBoxCenter[0]+int(landmarksBoxSize[0]/2) > originalImageSize[1]:
        tmp = int(np.ceil(landmarksBoxCenter[0] + int(landmarksBoxSize[0]/2) - originalImageSize[1]))

        if image.ndim == 2:
            padWidth = ( (0,0), (0, tmp) )
        elif image.ndim == 3:
            padWidth = ( (0,0), (0, tmp), (0, 0) )

        image = np.pad(image, pad_width=padWidth, mode='edge' )

    # Same on top and bottom
    if landmarksBoxCenter[1]-int(landmarksBoxSize[1]/2) < 0:
        tmp = int(np.ceil( int(landmarksBoxSize[1]/2) - landmarksBoxCenter[1]))

        if image.ndim == 2:
            padWidth = ( (tmp, 0),(0,0) )
        elif image.ndim == 3:
            padWidth = ( (tmp, 0),(0,0), (0, 0) )

        image = np.pad(image, pad_width=padWidth, mode='edge' )
        landmarksBoxCenter = (landmarksBoxCenter[0], landmarksBoxCenter[1] + tmp)
        for l in range(0, coords.shape[1]):
            coords[0][l][1] = coords[0][l][1] + tmp

    if landmarksBoxCenter[1]+landmarksBoxSize[1]/2 > originalImageSize[0]:
        tmp = int(np.ceil(landmarksBoxCenter[1] + landmarksBoxSize[1]/2 - originalImageSize[0]))

        if image.ndim == 2:
            padWidth = ( (0, tmp),(0,0) )
        elif image.ndim == 3:
            padWidth = ( (0, tmp),(0,0), (0, 0) )

        image = np.pad(image, pad_width=padWidth, mode='edge' )

    ## Cutting and resizing
    for l in range(0, coords.shape[1]):
        # Moving/Cutting
        coords[0][l][0] = coords[0][l][0] - int(landmarksBoxCenter[0]-int(landmarksBoxSize[0]/2.0))
        coords[0][l][1] = coords[0][l][1] - int(landmarksBoxCenter[1]-int(landmarksBoxSize[1]/2.0))


        # Resizing
        coords[0][l][0] = coords[0][l][0] * cnnImageSize[1]/landmarksBoxSize[1]
        coords[0][l][1] = coords[0][l][1] * cnnImageSize[0]/landmarksBoxSize[0]

    image = np.reshape(
        imresize(
            image[int(landmarksBoxCenter[1]-int(landmarksBoxSize[1]/2.0)):int(landmarksBoxCenter[1]+int(landmarksBoxSize[1]/2.0)),
                  int(landmarksBoxCenter[0]-int(landmarksBoxSize[0]/2.0)):int(landmarksBoxCenter[0]+int(landmarksBoxSize[0]/2.0)),
                  :],
            cnnImageSize),
        [1,cnnImageSize[0],cnnImageSize[1], cnnImageSize[2]]
    )

    return image, coords

def registerAndCutImage(inputImage, inputLandmarks, referenceLandmarks, cnnImageSize):

    ## Calculating the transformation
    # T, R, t = registerPointClouds2D(np.reshape(inputLandmarks,[44,2]), np.reshape(referenceLandmarks,[44,2]))
    T, R, t = registerPointClouds2D(np.reshape(inputLandmarks[0,0:33,:],[33,2]), np.reshape(referenceLandmarks[0,0:33,:],[33,2]))

    invR = np.linalg.inv(R)
    t = np.reshape(t,[2,1])

    ## Transforming the landmarks
    coords = np.array(inputLandmarks)
    for i in range(coords.shape[1]):
        coords[0,i,:] = (np.dot( R, np.reshape(inputLandmarks[0,i,:], [2,1])) + t).ravel()


    ## Resamplig the image
    image = np.zeros([cnnImageSize[0], cnnImageSize[1], cnnImageSize[2]], dtype=inputImage.dtype)
    for x in range(cnnImageSize[1]):
        for y in range(cnnImageSize[0]):
            # [x,y] are the transformed coordinates
            transformedCoords = np.array([[x],[y]])

            notTransformedCoords = (np.dot( invR, transformedCoords - t)).ravel()

            for c in range(cnnImageSize[2]):
                if int(notTransformedCoords[0]) < inputImage.shape[1] and int(notTransformedCoords[1]) < inputImage.shape[0]:
                    if inputImage.ndim == 2:
                        image[y,x, c] = inputImage[int(notTransformedCoords[1]),int(notTransformedCoords[0])]
                    elif inputImage.ndim==3:
                        image[y,x, c] = inputImage[int(notTransformedCoords[1]),int(notTransformedCoords[0]), c]

    return image, coords

def registerPointClouds2D(A, B):
    '''
    Calculates the least-squares best-fit transform between corresponding 3D points A->B
    Input:
      A: Nx2 numpy array of corresponding 2D points
      B: Nx2 numpy array of corresponding 2D points
    Returns:
      T: 3x3 homogeneous transformation matrix
      R: 2x2 rotation matrix
      t: 2x1 column vector
    '''

    assert len(A) == len(B)


    zz = np.zeros(shape=[A.shape[0],1])
    A = np.append(A, zz, axis=1)
    B = np.append(B, zz, axis=1)


    # translate points to their centroids
    centroid_A = np.mean(A, axis=0)
    centroid_B = np.mean(B, axis=0)
    AA = A - centroid_A
    BB = B - centroid_B

    # rotation matrix
    H = np.dot(AA.T, BB)
    U, S, Vt = np.linalg.svd(H)
    R = np.dot(Vt.T, U.T)

    ##################################################
    trace = np.sum(S)
    f = 0.0
    for i in range(AA.shape[0]):
        f = f + np.linalg.norm(AA[i,:])**2

    if f != 0:
        s = trace/f
    else:
        s = 1.0

    R *= s
    ##################################################

    # translation
    t = centroid_B.T - np.dot(R,centroid_A.T)

    t = t[0:2]
    R = R[0:2,0:2]


    # homogeneous transformation
    T = np.identity(3)

    T[0:2, 0:2] = R
    T[0:2, 2] = t

    return T, R, t

def poly2mask(rowCoords, colCoords, imageShape):

    fill_row_coords, fill_col_coords = draw.polygon(rowCoords, colCoords, imageShape)

    mask = np.zeros(imageShape, dtype=np.bool)
    mask[fill_row_coords, fill_col_coords] = True

    return mask
